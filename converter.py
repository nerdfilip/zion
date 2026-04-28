import sys
import os
import html
import csv
import pandas as pd
from PyQt5.QtWidgets import (QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, 
                             QTextEdit, QPushButton, QFrame, QFileDialog)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QTimer

# ==========================================
# 1. BACKGROUND WORKER THREAD
# ==========================================
class ConverterWorker(QThread):
    # Signals to communicate with the main GUI thread
    update_status = pyqtSignal(str)
    finished = pyqtSignal()

    GERMAN_CHAR_MAP = {
        'ä': 'ae',
        'ö': 'oe',
        'ü': 'ue',
        'Ä': 'ae',
        'Ö': 'oe',
        'Ü': 'ue',
        'ß': 'ss',
    }

    EXCEL_EXTENSIONS = ('.xls', '.xlsx', '.xlsm', '.xlsb', '.xltx', '.xltm', '.ods')
    STREAMABLE_EXCEL_EXTENSIONS = ('.xlsx', '.xlsm', '.xltx', '.xltm')
    LARGE_FILE_THRESHOLD = 300 * 1024 * 1024  # 300MB

    def __init__(self, file_paths, output_dir, use_german_dict=False):
        super().__init__()
        self.file_paths = file_paths
        self.output_dir = output_dir
        self.use_german_dict = use_german_dict

    @classmethod
    def _replace_german_chars(cls, value):
        text = str(value)
        return ''.join(cls.GERMAN_CHAR_MAP.get(ch, ch) for ch in text)

    @classmethod
    def _normalize_header(cls, header):
        cleaned = cls._replace_german_chars(header).strip().lower()
        cleaned = pd.Series([cleaned]).str.replace(r'[^a-zA-Z0-9_]', '_', regex=True).iloc[0]
        cleaned = pd.Series([cleaned]).str.replace(r'_+', '_', regex=True).str.strip('_').iloc[0]
        if cleaned and cleaned[0].isdigit():
            cleaned = '_' + cleaned
        return cleaned or 'column'

    @classmethod
    def _sanitize_headers(cls, columns):
        seen = {}
        sanitized = []
        for col in columns:
            base = cls._normalize_header(col)
            count = seen.get(base, 0)
            if count == 0:
                candidate = base
            else:
                candidate = f"{base}_{count + 1}"
            seen[base] = count + 1
            sanitized.append(candidate)
        return sanitized

    def _apply_german_dictionary_headers(self, columns):
        return [self._replace_german_chars(col) for col in columns]

    @staticmethod
    def _normalize_header_preserve_style(header, fallback_index):
        if header is None:
            return f"Column {fallback_index}"
        text = str(header).strip()
        if not text:
            return f"Column {fallback_index}"
        return text

    @classmethod
    def _deduplicate_headers_preserve_style(cls, columns):
        seen = {}
        deduplicated = []
        for idx, col in enumerate(columns, start=1):
            base = cls._normalize_header_preserve_style(col, idx)
            count = seen.get(base, 0)
            if count == 0:
                candidate = base
            else:
                candidate = f"{base} ({count + 1})"
            seen[base] = count + 1
            deduplicated.append(candidate)
        return deduplicated

    def _apply_german_dictionary_records(self, df):
        for col in df.select_dtypes(include=['object']).columns:
            df[col] = df[col].apply(
                lambda v: self._replace_german_chars(v) if isinstance(v, str) else v
            )

    def _get_excel_engine(self, ext):
        engine_map = {
            '.xls': 'xlrd',
            '.xlsx': 'openpyxl',
            '.xlsm': 'openpyxl',
            '.xltx': 'openpyxl',
            '.xltm': 'openpyxl',
            '.xlsb': 'pyxlsb',
            '.ods': 'odf',
        }
        return engine_map.get(ext)

    def _transform_dataframe(self, df):
        if self.use_german_dict:
            # ON means transliterate only German characters and preserve everything else.
            df.columns = [
                self._replace_german_chars(col) if isinstance(col, str) else col
                for col in df.columns
            ]
        else:
            # OFF means keep headers and values unchanged.
            return df

        if self.use_german_dict:
            self._apply_german_dictionary_records(df)

        return df

    def _convert_excel_streaming(self, file_path, out_path):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "openpyxl is required for streaming large .xlsx/.xlsm files. Install with: pip install openpyxl"
            ) from exc

        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            worksheet = workbook.worksheets[0]
            rows = worksheet.iter_rows(values_only=True)

            first_row = next(rows, None)
            if first_row is None:
                with open(out_path, 'w', newline='', encoding='utf-8-sig') as csv_file:
                    csv_file.write('')
                return

            if self.use_german_dict:
                headers = [
                    self._replace_german_chars(value) if isinstance(value, str) else ("" if value is None else value)
                    for value in first_row
                ]
            else:
                # OFF means preserve headers exactly as they are in Excel.
                headers = ["" if value is None else value for value in first_row]

            with open(out_path, 'w', newline='', encoding='utf-8-sig') as csv_file:
                writer = csv.writer(csv_file)
                writer.writerow(headers)

                for row in rows:
                    values = []
                    for cell_value in row:
                        if cell_value is None:
                            values.append('')
                        elif self.use_german_dict and isinstance(cell_value, str):
                            values.append(self._replace_german_chars(cell_value))
                        else:
                            values.append(cell_value)
                    writer.writerow(values)
        finally:
            workbook.close()

    def _convert_excel_file(self, file_path, out_path):
        ext = os.path.splitext(file_path)[1].lower()
        file_size = os.path.getsize(file_path)

        if file_size > self.LARGE_FILE_THRESHOLD and ext in self.STREAMABLE_EXCEL_EXTENSIONS:
            self.update_status.emit(
                f"Warning: Large Excel file detected ({file_size / (1024 * 1024):.1f}MB). Using streaming conversion."
            )
            self._convert_excel_streaming(file_path, out_path)
            return

        if file_size > self.LARGE_FILE_THRESHOLD:
            self.update_status.emit(
                "Warning: Large Excel file detected, using pandas fallback (higher memory usage expected)."
            )

        engine = self._get_excel_engine(ext)
        df = pd.read_excel(file_path, engine=engine)
        transformed_df = self._transform_dataframe(df)
        transformed_df.to_csv(out_path, index=False, encoding='utf-8-sig')

    def run(self):
        for file in self.file_paths:
            try:
                self.update_status.emit(f"Processing: {os.path.basename(file)}...")

                ext = os.path.splitext(file)[1].lower()
                base_name = os.path.splitext(os.path.basename(file))[0]
                out_path = os.path.join(self.output_dir, base_name + '.csv')

                if ext in self.EXCEL_EXTENSIONS:
                    self._convert_excel_file(file, out_path)
                else:
                    self.update_status.emit(f"Skipped {os.path.basename(file)} (Only Excel formats are supported)")
                    continue

                self.update_status.emit(f"Success: Saved {os.path.basename(out_path)}")
            except Exception as e:
                self.update_status.emit(f"Error at {os.path.basename(file)}: {str(e)}")
        
        # Emit finished signal when the loop is done
        self.finished.emit()

# ==========================================
# 2. MAIN GUI APPLICATION
# ==========================================
class DataConverterApp(QWidget):
    def __init__(self):
        super().__init__()
        self.file_list = []
        self.output_dir = ""
        self.time_elapsed = 0
        self.initUI()

    def initUI(self):
        self.setWindowTitle('Excel to CSV Converter')
        self.resize(920, 700)
        self.setAcceptDrops(True)

        self.setStyleSheet("""
            QWidget {
                background-color: #f7fafc;
                color: #1f2933;
                font-family: 'Segoe UI';
            }
            QLabel#title {
                font-size: 26px;
                font-weight: 700;
                color: #16324f;
            }
            QLabel#subtitle {
                color: #486581;
                font-size: 14px;
            }
            QFrame#card {
                background: #ffffff;
                border: 1px solid #d8e3ef;
                border-radius: 14px;
            }
            QPushButton {
                background-color: #0f766e;
                color: #ffffff;
                font-size: 13px;
                padding: 10px 14px;
                border: none;
                border-radius: 8px;
                font-weight: 600;
            }
            QPushButton:hover {
                background-color: #0d5f59;
            }
            QPushButton:checked {
                background-color: #1d4ed8;
            }
            QPushButton:disabled {
                background-color: #bcccdc;
                color: #627d98;
            }
            QComboBox {
                border: 1px solid #bcccdc;
                border-radius: 8px;
                padding: 7px;
                background: #ffffff;
            }
            QTextEdit {
                border: 1px solid #243b53;
                border-radius: 12px;
                background-color: #0b243d;
                color: #ffffff;
                font-family: Consolas, monospace;
                font-size: 12px;
            }
        """)

        layout = QVBoxLayout()
        layout.setSpacing(12)

        hero_card = QFrame()
        hero_card.setObjectName("card")
        hero_layout = QVBoxLayout(hero_card)
        hero_layout.setContentsMargins(16, 14, 16, 14)
        hero_layout.setSpacing(4)

        title = QLabel("Universal Data Converter")
        title.setObjectName("title")
        subtitle = QLabel("Convert Excel files of all common types to CSV output")
        subtitle.setObjectName("subtitle")
        hero_layout.addWidget(title)
        hero_layout.addWidget(subtitle)
        layout.addWidget(hero_card)

        # Drag & Drop Zone
        self.drop_zone = QLabel("\nDrop Excel files here\n")
        self.drop_zone.setAlignment(Qt.AlignCenter)
        self.drop_zone.setStyleSheet("""
            QLabel {
                background-color: #ffffff;
                font-size: 16px;
                color: #486581;
                border: 2px dashed #88a4be;
                border-radius: 12px;
                padding: 18px;
            }
        """)
        layout.addWidget(self.drop_zone)

        self.file_count_label = QLabel("Files selected: 0")
        self.file_count_label.setStyleSheet("font-size: 13px; color: #334e68; font-weight: 600;")
        layout.addWidget(self.file_count_label)

        # Settings Section
        settings_card = QFrame()
        settings_card.setObjectName("card")
        settings_layout = QVBoxLayout(settings_card)
        settings_layout.setContentsMargins(14, 14, 14, 14)
        settings_layout.setSpacing(10)

        top_settings_row = QHBoxLayout()
        
        # Output Folder Selection
        self.folder_btn = QPushButton("Select Output Folder")
        self.folder_btn.clicked.connect(self.select_output_folder)
        self.folder_label = QLabel("Save to: Not selected")
        self.folder_label.setStyleSheet("font-size: 13px; color: #334e68;")
        
        top_settings_row.addWidget(self.folder_btn)
        top_settings_row.addWidget(self.folder_label, stretch=1)
        settings_layout.addLayout(top_settings_row)

        dict_row = QHBoxLayout()
        self.german_dict_btn = QPushButton("German Dictionary: OFF")
        self.german_dict_btn.setCheckable(True)
        self.german_dict_btn.toggled.connect(self.toggle_german_dictionary)
        dict_row.addWidget(self.german_dict_btn)
        dict_row.addStretch(1)
        settings_layout.addLayout(dict_row)

        autodetect_hint = QLabel("Supported: .xls, .xlsx, .xlsm, .xlsb, .xltx, .xltm, .ods -> .csv")
        autodetect_hint.setStyleSheet("font-size: 12px; color: #486581;")
        settings_layout.addWidget(autodetect_hint)

        large_file_hint = QLabel("Large Excel files over 300MB use a streaming mode when possible.")
        large_file_hint.setStyleSheet("font-size: 12px; color: #486581;")
        settings_layout.addWidget(large_file_hint)

        layout.addWidget(settings_card)

        # Buttons Layout
        btn_layout = QHBoxLayout()

        # Browse Button
        self.browse_btn = QPushButton("Browse Files")
        self.browse_btn.clicked.connect(self.browse_files)
        btn_layout.addWidget(self.browse_btn)

        # Start Button
        self.start_btn = QPushButton("Start Conversion")
        self.start_btn.setEnabled(False)
        self.start_btn.clicked.connect(self.start_conversion)
        btn_layout.addWidget(self.start_btn)

        layout.addLayout(btn_layout)

        # Timer Label
        self.timer_label = QLabel("Time elapsed: 00:00")
        self.timer_label.setAlignment(Qt.AlignCenter)
        self.timer_label.setStyleSheet("font-size: 15px; font-weight: bold; color: #9f1239;")
        layout.addWidget(self.timer_label)

        # Real-time Status Box
        self.status_box = QTextEdit()
        self.status_box.setReadOnly(True)
        layout.addWidget(self.status_box)

        self.setLayout(layout)
        
        # Initialize Timer
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_timer)

    # --- Drag & Drop Events ---
    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.accept()
            self.drop_zone.setStyleSheet("""
                QLabel { background-color: #e6fffa; font-size: 16px; color: #0f172a; border: 2px dashed #0f766e; border-radius: 12px; padding: 18px; }
            """)
        else:
            event.ignore()

    def dragLeaveEvent(self, event):
        self.drop_zone.setStyleSheet("""
            QLabel { background-color: #ffffff; font-size: 16px; color: #486581; border: 2px dashed #88a4be; border-radius: 12px; padding: 18px; }
        """)

    def dropEvent(self, event):
        self.drop_zone.setStyleSheet("""
            QLabel { background-color: #ffffff; font-size: 16px; color: #486581; border: 2px dashed #88a4be; border-radius: 12px; padding: 18px; }
        """)
        for url in event.mimeData().urls():
            file_path = url.toLocalFile()
            if file_path.lower().endswith(ConverterWorker.EXCEL_EXTENSIONS) and file_path not in self.file_list:
                self.file_list.append(file_path)
        self.check_ready_state()

    # --- File & Folder Selection ---
    def browse_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "Select Files",
            "",
            "Excel Files (*.xls *.xlsx *.xlsm *.xlsb *.xltx *.xltm *.ods);;All Files (*)"
        )
        if files:
            for file in files:
                if file not in self.file_list:
                    self.file_list.append(file)
            self.check_ready_state()

    def select_output_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Output Folder")
        if folder:
            self.output_dir = folder
            self.folder_label.setText(f"Save to: {self.output_dir}")
            self.check_ready_state()

    # --- Validation ---
    def check_ready_state(self):
        self.file_count_label.setText(f"Files selected: {len(self.file_list)}")
        if self.file_list and self.output_dir:
            self.status_box.setText(
                f"{len(self.file_list)} file(s) ready.\n"
                "Output folder selected.\n"
                "Press 'Start Conversion' when ready.\n"
            )
            self.start_btn.setEnabled(True)
        elif self.file_list and not self.output_dir:
            self.status_box.setText(f"Added {len(self.file_list)} file(s).\nPlease select an output folder to continue.")
            self.start_btn.setEnabled(False)
        else:
            self.start_btn.setEnabled(False)

    def append_status(self, message):
        msg = str(message)
        lower = msg.lower()
        color = '#ffffff'
        if lower.startswith('success'):
            color = '#22c55e'
        elif lower.startswith('warning'):
            color = '#facc15'
        elif lower.startswith('error'):
            color = '#ef4444'

        safe_text = html.escape(msg)
        self.status_box.append(f"<span style='color: {color};'>{safe_text}</span>")

    def toggle_german_dictionary(self, enabled):
        if enabled:
            self.german_dict_btn.setText("German Dictionary: ON")
            self.append_status("Success: German dictionary enabled for headers and records")
        else:
            self.german_dict_btn.setText("German Dictionary: OFF")
            self.append_status("Warning: German dictionary disabled")

    # --- Conversion Logic ---
    def start_conversion(self):
        self.start_btn.setEnabled(False)
        self.browse_btn.setEnabled(False)
        self.folder_btn.setEnabled(False)

        german_enabled = self.german_dict_btn.isChecked()

        self.time_elapsed = 0
        self.timer_label.setText("Time elapsed: 00:00")
        self.append_status("-" * 40)
        self.append_status("Starting conversion to CSV format")
        self.append_status(f"Output directory: {self.output_dir}")
        if german_enabled:
            self.append_status("German dictionary: enabled (headers and records)")
        else:
            self.append_status("German dictionary: disabled")
        
        self.timer.start(1000)

        self.worker = ConverterWorker(
            self.file_list,
            self.output_dir,
            german_enabled
        )
        self.worker.update_status.connect(self.log_status)
        self.worker.finished.connect(self.conversion_finished)
        self.worker.start()

    def update_timer(self):
        self.time_elapsed += 1
        minutes = self.time_elapsed // 60
        seconds = self.time_elapsed % 60
        self.timer_label.setText(f"Time elapsed: {minutes:02d}:{seconds:02d}")

    def log_status(self, message):
        self.append_status(message)
        scrollbar = self.status_box.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())

    def conversion_finished(self):
        self.timer.stop()
        self.append_status("-" * 40)
        self.append_status("Success: Conversion finished.")
        
        self.file_list = [] 
        self.browse_btn.setEnabled(True)
        self.folder_btn.setEnabled(True)
        self.check_ready_state()

# ==========================================
# 3. RUN APPLICATION
# ==========================================
if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = DataConverterApp()
    window.show()
    sys.exit(app.exec_())